"""Shared XLSX export service for Products, Categories, and Reporting Groups (Stage 19).

Builds two kinds of workbook, both re-importable via import_service.py:
  - Template: header row + one example row + data-validation dropdowns for any
    foreign-key-ish column (category name, reporting group name).
  - Full export: the brand's current data for that entity.

Column headers are the shared contract with import_service.py — a header must
match exactly (case-insensitive) for that column to be recognised on import.
"""

import uuid
from datetime import date
from decimal import ROUND_HALF_UP, Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.category import Category
from app.models.product import Product
from app.models.product_attribute_type import ProductAttributeType
from app.models.product_attribute_value import ProductAttributeValue
from app.models.product_combo_group import ProductComboGroup
from app.models.product_variant import ProductVariant
from app.models.product_variant_attribute import ProductVariantAttribute
from app.models.reporting_group import ReportingGroup
from app.services.invoice_report_service import fetch_invoice_report_rows_for_export

# Header order is the on-disk column order for both template and full export.
PRODUCT_COLUMNS: list[str] = [
    "ref",
    "name",
    "category",
    "description",
    "print_name",
    "price",
    "is_taxable",
    "is_open_item",
    "display_order",
    "is_active",
]
CATEGORY_COLUMNS: list[str] = ["ref", "name", "reporting_group", "display_order", "is_active"]
REPORTING_GROUP_COLUMNS: list[str] = ["ref", "name"]
# "product_ref" (not product name) keys the linked product — refs are guaranteed
# unique per brand, unlike product names, so there's no ambiguity on import.
VARIANT_COLUMNS: list[str] = ["ref", "product_ref", "display_name", "attributes", "sku", "price", "is_active"]
COMBO_COLUMNS: list[str] = [
    "ref",
    "product_ref",
    "name",
    "display_name",
    "min_selections",
    "max_selections",
    "is_required",
    "display_order",
    "is_active",
]
INVOICE_COLUMNS: list[str] = [
    "id",
    "site",
    "invoice_type",
    "status",
    "created_at",
    "subtotal",
    "tax",
    "discount",
    "total",
    "is_refunded",
    "voided_at",
    "paid_at",
]

_BOOL_LABELS = {True: "TRUE", False: "FALSE"}


def _cents_to_dollars_str(cents: int) -> str:
    """
    Format a cents integer as a plain decimal dollar string, e.g. 1500 -> "15.00".

    Args:
        cents: Amount in cents (BIGINT storage, CLAUDE.md rule 4).

    Returns:
        str: Dollar amount with exactly two decimal places.
    """
    dollars = (Decimal(cents) / 100).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return f"{dollars:.2f}"


def _rows_to_workbook(headers: list[str], data_rows: list[dict[str, Any]], sheet_title: str) -> Workbook:
    """
    Write a header row plus data rows to a single-sheet workbook.

    Args:
        headers: Column header strings, in order.
        data_rows: Each dict maps a header to its cell value for that row.
        sheet_title: Name of the worksheet.

    Returns:
        Workbook: An in-memory openpyxl workbook, ready for data validation or streaming.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(headers)
    ws.freeze_panes = "A2"
    for row in data_rows:
        ws.append([row.get(h, "") for h in headers])
    return wb


def _add_name_dropdown(wb: Workbook, ws_title: str, column_index: int, options: list[str], max_rows: int = 1000) -> None:
    """
    Attach an Excel data-validation dropdown of existing names to a column.

    Options are written to a hidden helper sheet and referenced by range,
    since a literal inline list is capped at 255 characters by Excel and
    brands may have far more categories/reporting groups than that allows.

    Args:
        wb: The workbook to modify.
        ws_title: Title of the sheet the dropdown applies to.
        column_index: 1-based column index the dropdown is attached to.
        options: Valid option strings (e.g. existing category names).
        max_rows: How many data rows (below the header) get the dropdown applied.
    """
    if not options:
        return

    helper_title = f"_{ws_title}_lists"[:31]
    helper = wb.create_sheet(title=helper_title)
    for i, option in enumerate(options, start=1):
        helper.cell(row=i, column=column_index, value=option)
    helper.sheet_state = "hidden"

    from openpyxl.utils import get_column_letter

    col_letter = get_column_letter(column_index)
    formula = f"{helper_title}!${col_letter}$1:${col_letter}${len(options)}"
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showErrorMessage=True)
    dv.error = "Select a value from the dropdown list"
    dv.errorTitle = "Invalid value"

    ws = wb[ws_title]
    dv.add(f"{col_letter}2:{col_letter}{max_rows}")
    ws.add_data_validation(dv)


def workbook_to_bytes(wb: Workbook) -> bytes:
    """
    Serialise a workbook to raw XLSX bytes for streaming in an HTTP response.

    Args:
        wb: The workbook to serialise.

    Returns:
        bytes: Raw .xlsx file content.
    """
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


async def _category_names(db: AsyncSession, brand_id: uuid.UUID) -> list[str]:
    """Return active category names for a brand, alphabetically, for dropdown options."""
    result = await db.execute(
        select(Category.name)
        .where(Category.brand_id == brand_id, Category.is_active == True)  # noqa: E712
        .order_by(Category.name)
    )
    return [row[0] for row in result.all()]


async def _reporting_group_names(db: AsyncSession, brand_id: uuid.UUID) -> list[str]:
    """Return reporting group names for a brand, alphabetically, for dropdown options."""
    result = await db.execute(
        select(ReportingGroup.name).where(ReportingGroup.brand_id == brand_id).order_by(ReportingGroup.name)
    )
    return [row[0] for row in result.all()]


async def _product_refs(db: AsyncSession, brand_id: uuid.UUID) -> list[str]:
    """Return active product ref codes for a brand, alphabetically, for dropdown options."""
    result = await db.execute(
        select(Product.ref)
        .where(Product.brand_id == brand_id, Product.is_active == True)  # noqa: E712
        .order_by(Product.ref)
    )
    return [row[0] for row in result.all()]


async def _describe_variant_attributes(db: AsyncSession, variant_id: uuid.UUID) -> str:
    """
    Build a read-only "Type: Value, Type2: Value2" summary of a variant's attributes.

    Informational only — attribute assignment isn't part of the reimportable
    column set (VARIANT_COLUMNS), since it varies per brand and doesn't fit a
    fixed spreadsheet header; import_service.py ignores this column.
    """
    result = await db.execute(
        select(ProductAttributeType.name, ProductAttributeValue.value)
        .select_from(ProductVariantAttribute)
        .join(ProductAttributeType, ProductVariantAttribute.attribute_type_id == ProductAttributeType.id)
        .join(ProductAttributeValue, ProductVariantAttribute.attribute_value_id == ProductAttributeValue.id)
        .where(ProductVariantAttribute.variant_id == variant_id)
    )
    return ", ".join(f"{type_name}: {value}" for type_name, value in result.all())


# ── Products ──────────────────────────────────────────────────────────────────


async def export_products(db: AsyncSession, brand_id: uuid.UUID) -> list[dict[str, Any]]:
    """
    Fetch all active products for a brand as export-ready rows keyed by PRODUCT_COLUMNS.

    Args:
        db: Active database session.
        brand_id: Brand to export products for.

    Returns:
        list[dict]: One dict per product, ordered by display_order then name.
    """
    result = await db.execute(
        select(Product, Category.name)
        .join(Category, Product.category_id == Category.id)
        .where(Product.brand_id == brand_id, Product.is_active == True)  # noqa: E712
        .order_by(Product.display_order, Product.name)
    )
    rows: list[dict[str, Any]] = []
    for product, category_name in result.all():
        rows.append(
            {
                "ref": product.ref,
                "name": product.name,
                "category": category_name,
                "description": product.description or "",
                "print_name": product.print_name or "",
                "price": _cents_to_dollars_str(product.base_price_cents),
                "is_taxable": _BOOL_LABELS[product.is_taxable],
                "is_open_item": _BOOL_LABELS[product.is_open_item],
                "display_order": product.display_order,
                "is_active": _BOOL_LABELS[product.is_active],
            }
        )
    return rows


async def build_products_template(db: AsyncSession, brand_id: uuid.UUID) -> Workbook:
    """
    Build a Products import template: header + one example row + category dropdown.

    Args:
        db: Active database session.
        brand_id: Brand the template's dropdown options are scoped to.

    Returns:
        Workbook: Ready to stream as an .xlsx download.
    """
    category_names = await _category_names(db, brand_id)
    example = {
        "ref": "",
        "name": "Example Product",
        "category": category_names[0] if category_names else "Uncategorised",
        "description": "",
        "print_name": "",
        "price": "9.99",
        "is_taxable": "TRUE",
        "is_open_item": "FALSE",
        "display_order": "0",
        "is_active": "TRUE",
    }
    wb = _rows_to_workbook(PRODUCT_COLUMNS, [example], "Products")
    _add_name_dropdown(wb, "Products", PRODUCT_COLUMNS.index("category") + 1, category_names)
    return wb


async def build_products_export(db: AsyncSession, brand_id: uuid.UUID) -> Workbook:
    """
    Build a full Products export workbook with the brand's current catalog data.

    Args:
        db: Active database session.
        brand_id: Brand to export.

    Returns:
        Workbook: Ready to stream as an .xlsx download.
    """
    rows = await export_products(db, brand_id)
    category_names = await _category_names(db, brand_id)
    wb = _rows_to_workbook(PRODUCT_COLUMNS, rows, "Products")
    _add_name_dropdown(wb, "Products", PRODUCT_COLUMNS.index("category") + 1, category_names, max_rows=max(len(rows) + 500, 1000))
    return wb


# ── Categories ────────────────────────────────────────────────────────────────


async def export_categories(db: AsyncSession, brand_id: uuid.UUID) -> list[dict[str, Any]]:
    """
    Fetch all categories for a brand as export-ready rows keyed by CATEGORY_COLUMNS.

    Args:
        db: Active database session.
        brand_id: Brand to export categories for.

    Returns:
        list[dict]: One dict per category, ordered by display_order then name.
    """
    result = await db.execute(
        select(Category, ReportingGroup.name)
        .join(ReportingGroup, Category.reporting_group_id == ReportingGroup.id)
        .where(Category.brand_id == brand_id)
        .order_by(Category.display_order, Category.name)
    )
    rows: list[dict[str, Any]] = []
    for category, reporting_group_name in result.all():
        rows.append(
            {
                "ref": category.ref,
                "name": category.name,
                "reporting_group": reporting_group_name,
                "display_order": category.display_order,
                "is_active": _BOOL_LABELS[category.is_active],
            }
        )
    return rows


async def build_categories_template(db: AsyncSession, brand_id: uuid.UUID) -> Workbook:
    """
    Build a Categories import template: header + one example row + reporting group dropdown.

    Args:
        db: Active database session.
        brand_id: Brand the template's dropdown options are scoped to.

    Returns:
        Workbook: Ready to stream as an .xlsx download.
    """
    reporting_group_names = await _reporting_group_names(db, brand_id)
    example = {
        "ref": "",
        "name": "Example Category",
        "reporting_group": reporting_group_names[0] if reporting_group_names else "",
        "display_order": "0",
        "is_active": "TRUE",
    }
    wb = _rows_to_workbook(CATEGORY_COLUMNS, [example], "Categories")
    _add_name_dropdown(wb, "Categories", CATEGORY_COLUMNS.index("reporting_group") + 1, reporting_group_names)
    return wb


async def build_categories_export(db: AsyncSession, brand_id: uuid.UUID) -> Workbook:
    """
    Build a full Categories export workbook with the brand's current categories.

    Args:
        db: Active database session.
        brand_id: Brand to export.

    Returns:
        Workbook: Ready to stream as an .xlsx download.
    """
    rows = await export_categories(db, brand_id)
    reporting_group_names = await _reporting_group_names(db, brand_id)
    wb = _rows_to_workbook(CATEGORY_COLUMNS, rows, "Categories")
    _add_name_dropdown(
        wb,
        "Categories",
        CATEGORY_COLUMNS.index("reporting_group") + 1,
        reporting_group_names,
        max_rows=max(len(rows) + 500, 1000),
    )
    return wb


# ── Reporting Groups ──────────────────────────────────────────────────────────


async def export_reporting_groups(db: AsyncSession, brand_id: uuid.UUID) -> list[dict[str, Any]]:
    """
    Fetch all reporting groups for a brand as export-ready rows keyed by REPORTING_GROUP_COLUMNS.

    Args:
        db: Active database session.
        brand_id: Brand to export reporting groups for.

    Returns:
        list[dict]: One dict per reporting group, default group first.
    """
    result = await db.execute(
        select(ReportingGroup)
        .where(ReportingGroup.brand_id == brand_id)
        .order_by(ReportingGroup.is_default.desc(), ReportingGroup.name)
    )
    return [{"ref": group.ref, "name": group.name} for group in result.scalars().all()]


async def build_reporting_groups_template() -> Workbook:
    """
    Build a Reporting Groups import template: header + one example row (no FK columns).

    Returns:
        Workbook: Ready to stream as an .xlsx download.
    """
    example = {"ref": "", "name": "Example Reporting Group"}
    return _rows_to_workbook(REPORTING_GROUP_COLUMNS, [example], "Reporting Groups")


async def build_reporting_groups_export(db: AsyncSession, brand_id: uuid.UUID) -> Workbook:
    """
    Build a full Reporting Groups export workbook with the brand's current groups.

    Args:
        db: Active database session.
        brand_id: Brand to export.

    Returns:
        Workbook: Ready to stream as an .xlsx download.
    """
    rows = await export_reporting_groups(db, brand_id)
    return _rows_to_workbook(REPORTING_GROUP_COLUMNS, rows, "Reporting Groups")


# ── Variants (Stage 22) ───────────────────────────────────────────────────────
#
# Update-only: import_variants() only matches existing rows by ref, since
# creating a variant requires attribute assignment (brand-specific, doesn't
# fit a fixed spreadsheet header) — see import_service.py.


async def export_variants(db: AsyncSession, brand_id: uuid.UUID) -> list[dict[str, Any]]:
    """
    Fetch all active variants for a brand as export-ready rows keyed by VARIANT_COLUMNS.

    Args:
        db: Active database session.
        brand_id: Brand to export variants for.

    Returns:
        list[dict]: One dict per variant, ordered by product name.
    """
    result = await db.execute(
        select(ProductVariant, Product.ref)
        .join(Product, ProductVariant.product_id == Product.id)
        .where(Product.brand_id == brand_id, ProductVariant.is_active == True)  # noqa: E712
        .order_by(Product.name, ProductVariant.created_at)
    )
    rows: list[dict[str, Any]] = []
    for variant, product_ref in result.all():
        rows.append(
            {
                "ref": variant.ref,
                "product_ref": product_ref,
                "display_name": variant.display_name or "",
                "attributes": await _describe_variant_attributes(db, variant.id),
                "sku": variant.sku or "",
                "price": _cents_to_dollars_str(variant.price_cents) if variant.price_cents is not None else "",
                "is_active": _BOOL_LABELS[variant.is_active],
            }
        )
    return rows


async def build_variants_template(db: AsyncSession, brand_id: uuid.UUID) -> Workbook:
    """
    Build a Variants import template: header + one example row + product dropdown.

    Args:
        db: Active database session.
        brand_id: Brand the template's dropdown options are scoped to.

    Returns:
        Workbook: Ready to stream as an .xlsx download.
    """
    product_refs = await _product_refs(db, brand_id)
    example = {
        "ref": "",
        "product_ref": product_refs[0] if product_refs else "",
        "display_name": "Example Variant",
        "attributes": "",
        "sku": "",
        "price": "9.99",
        "is_active": "TRUE",
    }
    wb = _rows_to_workbook(VARIANT_COLUMNS, [example], "Variants")
    _add_name_dropdown(wb, "Variants", VARIANT_COLUMNS.index("product_ref") + 1, product_refs)
    return wb


async def build_variants_export(db: AsyncSession, brand_id: uuid.UUID) -> Workbook:
    """
    Build a full Variants export workbook with the brand's current variants.

    Args:
        db: Active database session.
        brand_id: Brand to export.

    Returns:
        Workbook: Ready to stream as an .xlsx download.
    """
    rows = await export_variants(db, brand_id)
    product_refs = await _product_refs(db, brand_id)
    wb = _rows_to_workbook(VARIANT_COLUMNS, rows, "Variants")
    _add_name_dropdown(
        wb,
        "Variants",
        VARIANT_COLUMNS.index("product_ref") + 1,
        product_refs,
        max_rows=max(len(rows) + 500, 1000),
    )
    return wb


# ── Combos (Stage 22) ─────────────────────────────────────────────────────────


async def export_combos(db: AsyncSession, brand_id: uuid.UUID) -> list[dict[str, Any]]:
    """
    Fetch all active combo groups for a brand as export-ready rows keyed by COMBO_COLUMNS.

    Args:
        db: Active database session.
        brand_id: Brand to export combo groups for.

    Returns:
        list[dict]: One dict per combo group, ordered by product name.
    """
    result = await db.execute(
        select(ProductComboGroup, Product.ref)
        .join(Product, ProductComboGroup.product_id == Product.id)
        .where(Product.brand_id == brand_id, ProductComboGroup.is_active == True)  # noqa: E712
        .order_by(Product.name, ProductComboGroup.display_order)
    )
    rows: list[dict[str, Any]] = []
    for group, product_ref in result.all():
        rows.append(
            {
                "ref": group.ref,
                "product_ref": product_ref,
                "name": group.name,
                "display_name": group.display_name or "",
                "min_selections": group.min_selections,
                "max_selections": group.max_selections,
                "is_required": _BOOL_LABELS[group.is_required],
                "display_order": group.display_order,
                "is_active": _BOOL_LABELS[group.is_active],
            }
        )
    return rows


async def build_combos_template(db: AsyncSession, brand_id: uuid.UUID) -> Workbook:
    """
    Build a Combos import template: header + one example row + product dropdown.

    Args:
        db: Active database session.
        brand_id: Brand the template's dropdown options are scoped to.

    Returns:
        Workbook: Ready to stream as an .xlsx download.
    """
    product_refs = await _product_refs(db, brand_id)
    example = {
        "ref": "",
        "product_ref": product_refs[0] if product_refs else "",
        "name": "Choose a side",
        "display_name": "Example Combo",
        "min_selections": "1",
        "max_selections": "1",
        "is_required": "TRUE",
        "display_order": "0",
        "is_active": "TRUE",
    }
    wb = _rows_to_workbook(COMBO_COLUMNS, [example], "Combos")
    _add_name_dropdown(wb, "Combos", COMBO_COLUMNS.index("product_ref") + 1, product_refs)
    return wb


async def build_combos_export(db: AsyncSession, brand_id: uuid.UUID) -> Workbook:
    """
    Build a full Combos export workbook with the brand's current combo groups.

    Args:
        db: Active database session.
        brand_id: Brand to export.

    Returns:
        Workbook: Ready to stream as an .xlsx download.
    """
    rows = await export_combos(db, brand_id)
    product_refs = await _product_refs(db, brand_id)
    wb = _rows_to_workbook(COMBO_COLUMNS, rows, "Combos")
    _add_name_dropdown(
        wb,
        "Combos",
        COMBO_COLUMNS.index("product_ref") + 1,
        product_refs,
        max_rows=max(len(rows) + 500, 1000),
    )
    return wb


# ── Invoices (Stage 21) ───────────────────────────────────────────────────────
#
# Read-only export — there is no matching import_invoices(); invoices are
# created by the sale flow, not bulk-uploaded, so this reuses only the
# workbook-building half of the Stage 19 framework.


async def export_invoices(
    db: AsyncSession,
    brand_id: uuid.UUID,
    site_id: uuid.UUID | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    invoice_status: str | None = None,
    min_amount_cents: int | None = None,
    max_amount_cents: int | None = None,
) -> list[dict[str, Any]]:
    """
    Fetch filtered invoices for a brand as export-ready rows keyed by INVOICE_COLUMNS.

    Filters mirror the invoice reporting list endpoint so "export the filtered
    set" produces exactly what's on screen.

    Args:
        db: Active database session.
        brand_id: Brand to export invoices for.
        site_id: Optional site filter.
        start_date: Optional lower bound on created_at date (inclusive).
        end_date: Optional upper bound on created_at date (inclusive).
        invoice_status: Optional invoice status filter.
        min_amount_cents: Optional lower bound on total_cents.
        max_amount_cents: Optional upper bound on total_cents.

    Returns:
        list[dict]: One dict per invoice, ordered most-recent-first.
    """
    rows = await fetch_invoice_report_rows_for_export(
        db, brand_id, site_id, start_date, end_date, invoice_status, min_amount_cents, max_amount_cents
    )
    return [
        {
            "id": str(row["id"]),
            "site": row["site_name"],
            "invoice_type": row["invoice_type"],
            "status": row["status"],
            "created_at": row["created_at"].isoformat(),
            "subtotal": _cents_to_dollars_str(row["subtotal_cents"]),
            "tax": _cents_to_dollars_str(row["tax_cents"]),
            "discount": _cents_to_dollars_str(row["discount_cents"]),
            "total": _cents_to_dollars_str(row["total_cents"]),
            "is_refunded": _BOOL_LABELS[row["is_refunded"]],
            "voided_at": row["voided_at"].isoformat() if row["voided_at"] else "",
            "paid_at": row["paid_at"].isoformat() if row["paid_at"] else "",
        }
        for row in rows
    ]


async def build_invoices_export(
    db: AsyncSession,
    brand_id: uuid.UUID,
    site_id: uuid.UUID | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    invoice_status: str | None = None,
    min_amount_cents: int | None = None,
    max_amount_cents: int | None = None,
) -> Workbook:
    """
    Build a filtered Invoices export workbook.

    Args:
        db: Active database session.
        brand_id: Brand to export.
        site_id: Optional site filter.
        start_date: Optional lower bound on created_at date (inclusive).
        end_date: Optional upper bound on created_at date (inclusive).
        invoice_status: Optional invoice status filter.
        min_amount_cents: Optional lower bound on total_cents.
        max_amount_cents: Optional upper bound on total_cents.

    Returns:
        Workbook: Ready to stream as an .xlsx download.
    """
    rows = await export_invoices(
        db, brand_id, site_id, start_date, end_date, invoice_status, min_amount_cents, max_amount_cents
    )
    return _rows_to_workbook(INVOICE_COLUMNS, rows, "Invoices")

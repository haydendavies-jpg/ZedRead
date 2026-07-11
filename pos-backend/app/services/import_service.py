"""Shared XLSX import service for Products, Categories, and Reporting Groups (Stage 19).

Validate-then-upsert semantics: each row is validated independently and only
applied if it passes. A row matched to an existing record by its `ref` column
is updated; only the columns actually present in the uploaded header row are
touched (partial-update semantics) — columns from the schema that aren't in
the header are left untouched. A row with no `ref` is treated as a new record
and assigned the next ref from that entity's sequence, same as a normal create.

A row that fails validation (unknown ref, unresolvable category/reporting
group name, bad type) is skipped and reported in ImportSummary.errors rather
than aborting the whole upload — so one bad row doesn't block the rest of a
large sheet.

All rows from one call share a single `import_id` (a fresh UUID per upload),
threaded through to the underlying create/update service functions so every
audit_logs row they write carries it in after_state — that's what lets a
whole import be traced/reasoned about together (no new table needed).
"""

import uuid
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.category import Category
from app.models.product import Product
from app.models.product_combo_group import ProductComboGroup
from app.models.product_variant import ProductVariant
from app.models.reporting_group import ReportingGroup
from app.models.superadmin import SuperAdmin
from app.models.user import User
from app.schemas.category import CategoryCreate, CategoryUpdate
from app.schemas.combo import ComboGroupCreate, ComboGroupUpdate
from app.schemas.import_export import ImportRowError, ImportSummary
from app.schemas.product import ProductCreate, ProductUpdate
from app.schemas.reporting_group import ReportingGroupCreate, ReportingGroupUpdate
from app.schemas.variant import VariantUpdate
from app.services import category_service, combo_service, product_service, reporting_group_service, variant_service


class InvalidWorkbookError(ValueError):
    """Raised when the uploaded file cannot be parsed as an XLSX workbook."""


def parse_xlsx(file_bytes: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    """
    Parse raw XLSX bytes into a lowercased header list and a list of row dicts.

    The first non-empty row is treated as the header row. Each subsequent row
    is returned as a dict mapping header name -> raw cell value, skipping
    fully-blank rows. Unrecognised header cells (blank) are dropped so callers
    only see named columns.

    Args:
        file_bytes: Raw bytes of the uploaded .xlsx file.

    Returns:
        tuple[list[str], list[dict[str, Any]]]: (headers, rows).

    Raises:
        InvalidWorkbookError: If the bytes cannot be parsed as an .xlsx file.
    """
    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
        ws = wb.worksheets[0]
        rows_iter = ws.iter_rows(values_only=True)
    except Exception as exc:
        raise InvalidWorkbookError("File is not a valid .xlsx workbook") from exc

    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []

    raw_headers = [str(h).strip().lower() if h is not None else "" for h in header_row]

    data_rows: list[dict[str, Any]] = []
    for raw_row in rows_iter:
        if raw_row is None or all(cell is None for cell in raw_row):
            continue
        row_dict = {
            raw_headers[i]: raw_row[i]
            for i in range(min(len(raw_headers), len(raw_row)))
            if raw_headers[i]
        }
        if row_dict:
            data_rows.append(row_dict)

    headers = [h for h in raw_headers if h]
    return headers, data_rows


# ── Value coercion helpers ────────────────────────────────────────────────────


def _clean_str(value: Any) -> str | None:
    """Return a trimmed string, or None if the cell is blank/absent."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_bool(value: Any) -> bool | None:
    """
    Parse a cell as a boolean. Accepts TRUE/FALSE, 1/0, yes/no (case-insensitive).

    Raises:
        ValueError: If the cell has content that isn't a recognised boolean.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text == "":
        return None
    if text in ("true", "1", "yes", "y"):
        return True
    if text in ("false", "0", "no", "n"):
        return False
    raise ValueError(f"Invalid boolean value '{value}'")


def _parse_int(value: Any) -> int | None:
    """Parse a cell as an integer, or None if blank."""
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    try:
        return int(float(text))
    except ValueError as exc:
        raise ValueError(f"Invalid integer value '{value}'") from exc


def _parse_price_cents(value: Any) -> int | None:
    """Parse a dollar-amount cell (e.g. '9.99') into cents (BIGINT storage, rule 4/9)."""
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    try:
        dollars = Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Invalid price value '{value}'") from exc
    return int((dollars * 100).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _error_message(exc: Exception) -> str:
    """Extract a human-readable message from a ValueError or HTTPException."""
    if isinstance(exc, HTTPException):
        return str(exc.detail)
    return str(exc)


async def _find_by_ref(db: AsyncSession, model: type, brand_id: uuid.UUID, ref: str) -> Any | None:
    """Look up a brand-scoped row by its human-readable ref code."""
    result = await db.execute(select(model).where(model.ref == ref, model.brand_id == brand_id))
    return result.scalar_one_or_none()


async def _category_id_by_name(db: AsyncSession, brand_id: uuid.UUID) -> dict[str, uuid.UUID]:
    """Map lowercased category name -> id for a brand (for the 'category' import column)."""
    result = await db.execute(select(Category).where(Category.brand_id == brand_id))
    return {c.name.strip().lower(): c.id for c in result.scalars().all()}


async def _reporting_group_id_by_name(db: AsyncSession, brand_id: uuid.UUID) -> dict[str, uuid.UUID]:
    """Map lowercased reporting group name -> id for a brand (for the 'reporting_group' column)."""
    result = await db.execute(select(ReportingGroup).where(ReportingGroup.brand_id == brand_id))
    return {g.name.strip().lower(): g.id for g in result.scalars().all()}


async def _product_id_by_ref(db: AsyncSession, brand_id: uuid.UUID) -> dict[str, uuid.UUID]:
    """Map product ref -> id for a brand (for the 'product_ref' column on Variants/Combos)."""
    result = await db.execute(select(Product).where(Product.brand_id == brand_id))
    return {p.ref: p.id for p in result.scalars().all()}


async def _find_variant_by_ref(db: AsyncSession, brand_id: uuid.UUID, ref: str) -> ProductVariant | None:
    """Look up a brand-scoped ProductVariant by ref, joining through its parent Product for scoping."""
    result = await db.execute(
        select(ProductVariant)
        .join(Product, ProductVariant.product_id == Product.id)
        .where(ProductVariant.ref == ref, Product.brand_id == brand_id)
    )
    return result.scalar_one_or_none()


async def _find_combo_group_by_ref(db: AsyncSession, brand_id: uuid.UUID, ref: str) -> ProductComboGroup | None:
    """Look up a brand-scoped ProductComboGroup by ref, joining through its parent Product for scoping."""
    result = await db.execute(
        select(ProductComboGroup)
        .join(Product, ProductComboGroup.product_id == Product.id)
        .where(ProductComboGroup.ref == ref, Product.brand_id == brand_id)
    )
    return result.scalar_one_or_none()


# ── Products ──────────────────────────────────────────────────────────────────


async def import_products(
    db: AsyncSession,
    brand_id: uuid.UUID,
    file_bytes: bytes,
    actor: User | SuperAdmin,
) -> ImportSummary:
    """
    Bulk import Products from an uploaded XLSX sheet (see PRODUCT_COLUMNS in export_service.py).

    Args:
        db: Active database session.
        brand_id: Brand to import into.
        file_bytes: Raw bytes of the uploaded .xlsx file.
        actor: The authenticated user performing the import (for audit logging).

    Returns:
        ImportSummary: Created/updated counts and any skipped-row errors.
    """
    headers, rows = parse_xlsx(file_bytes)
    import_id = uuid.uuid4()
    created = 0
    updated = 0
    errors: list[ImportRowError] = []

    category_map = await _category_id_by_name(db, brand_id) if "category" in headers else {}

    for row_number, row in enumerate(rows, start=2):
        try:
            ref = _clean_str(row.get("ref"))

            category_id: uuid.UUID | None = None
            if "category" in headers:
                category_name = _clean_str(row.get("category"))
                if category_name is not None:
                    category_id = category_map.get(category_name.lower())
                    if category_id is None:
                        raise ValueError(f"Unknown category '{category_name}'")

            name = _clean_str(row.get("name")) if "name" in headers else None
            description = _clean_str(row.get("description")) if "description" in headers else None
            print_name = _clean_str(row.get("print_name")) if "print_name" in headers else None
            price_cents = _parse_price_cents(row.get("price")) if "price" in headers else None
            is_taxable = _parse_bool(row.get("is_taxable")) if "is_taxable" in headers else None
            is_open_item = _parse_bool(row.get("is_open_item")) if "is_open_item" in headers else None
            display_order = _parse_int(row.get("display_order")) if "display_order" in headers else None
            is_active = _parse_bool(row.get("is_active")) if "is_active" in headers else None

            if ref:
                product = await _find_by_ref(db, Product, brand_id, ref)
                if product is None:
                    raise ValueError(f"No product found with ref '{ref}'")

                payload = ProductUpdate(
                    category_id=category_id,
                    name=name,
                    description=description,
                    print_name=print_name,
                    base_price_cents=price_cents,
                    is_taxable=is_taxable,
                    is_open_item=is_open_item,
                    display_order=display_order,
                )
                await product_service.update_product(
                    db, brand_id, product.id, payload, actor, import_id=import_id
                )
                if is_active is not None:
                    await product_service.set_product_active_state(
                        db, brand_id, product.id, is_active, actor, import_id=import_id
                    )
                updated += 1
            else:
                if not name or category_id is None or price_cents is None:
                    raise ValueError("New product rows require name, category, and price")

                payload = ProductCreate(
                    category_id=category_id,
                    name=name,
                    description=description,
                    print_name=print_name,
                    base_price_cents=price_cents,
                    is_taxable=is_taxable if is_taxable is not None else True,
                    is_open_item=is_open_item if is_open_item is not None else False,
                    display_order=display_order if display_order is not None else 0,
                )
                await product_service.create_product(db, brand_id, payload, actor, import_id=import_id)
                created += 1
        except (ValueError, HTTPException) as exc:
            errors.append(ImportRowError(row_number=row_number, message=_error_message(exc)))

    return ImportSummary(import_id=import_id, created=created, updated=updated, errors=errors)


# ── Categories ────────────────────────────────────────────────────────────────


async def import_categories(
    db: AsyncSession,
    brand_id: uuid.UUID,
    file_bytes: bytes,
    actor: User | SuperAdmin,
) -> ImportSummary:
    """
    Bulk import Categories from an uploaded XLSX sheet (see CATEGORY_COLUMNS in export_service.py).

    Args:
        db: Active database session.
        brand_id: Brand to import into.
        file_bytes: Raw bytes of the uploaded .xlsx file.
        actor: The authenticated user performing the import (for audit logging).

    Returns:
        ImportSummary: Created/updated counts and any skipped-row errors.
    """
    headers, rows = parse_xlsx(file_bytes)
    import_id = uuid.uuid4()
    created = 0
    updated = 0
    errors: list[ImportRowError] = []

    reporting_group_map = (
        await _reporting_group_id_by_name(db, brand_id) if "reporting_group" in headers else {}
    )

    for row_number, row in enumerate(rows, start=2):
        try:
            ref = _clean_str(row.get("ref"))

            reporting_group_id: uuid.UUID | None = None
            if "reporting_group" in headers:
                group_name = _clean_str(row.get("reporting_group"))
                if group_name is not None:
                    reporting_group_id = reporting_group_map.get(group_name.lower())
                    if reporting_group_id is None:
                        raise ValueError(f"Unknown reporting group '{group_name}'")

            name = _clean_str(row.get("name")) if "name" in headers else None
            display_order = _parse_int(row.get("display_order")) if "display_order" in headers else None
            is_active = _parse_bool(row.get("is_active")) if "is_active" in headers else None

            if ref:
                category = await _find_by_ref(db, Category, brand_id, ref)
                if category is None:
                    raise ValueError(f"No category found with ref '{ref}'")

                payload = CategoryUpdate(
                    name=name,
                    reporting_group_id=reporting_group_id,
                    display_order=display_order,
                    is_active=is_active,
                )
                await category_service.update_category(
                    db, brand_id, category.id, payload, actor, import_id=import_id
                )
                updated += 1
            else:
                if not name:
                    raise ValueError("New category rows require a name")

                payload = CategoryCreate(
                    name=name,
                    brand_id=brand_id,
                    reporting_group_id=reporting_group_id,
                    display_order=display_order if display_order is not None else 0,
                )
                await category_service.create_category(db, brand_id, payload, actor, import_id=import_id)
                created += 1
        except (ValueError, HTTPException) as exc:
            errors.append(ImportRowError(row_number=row_number, message=_error_message(exc)))

    return ImportSummary(import_id=import_id, created=created, updated=updated, errors=errors)


# ── Reporting Groups ──────────────────────────────────────────────────────────


async def import_reporting_groups(
    db: AsyncSession,
    brand_id: uuid.UUID,
    file_bytes: bytes,
    actor: User | SuperAdmin,
) -> ImportSummary:
    """
    Bulk import Reporting Groups from an uploaded XLSX sheet (REPORTING_GROUP_COLUMNS).

    Args:
        db: Active database session.
        brand_id: Brand to import into.
        file_bytes: Raw bytes of the uploaded .xlsx file.
        actor: The authenticated user performing the import (for audit logging).

    Returns:
        ImportSummary: Created/updated counts and any skipped-row errors.
    """
    headers, rows = parse_xlsx(file_bytes)
    import_id = uuid.uuid4()
    created = 0
    updated = 0
    errors: list[ImportRowError] = []

    for row_number, row in enumerate(rows, start=2):
        try:
            ref = _clean_str(row.get("ref"))
            name = _clean_str(row.get("name")) if "name" in headers else None

            if ref:
                group = await _find_by_ref(db, ReportingGroup, brand_id, ref)
                if group is None:
                    raise ValueError(f"No reporting group found with ref '{ref}'")

                payload = ReportingGroupUpdate(name=name)
                await reporting_group_service.update_reporting_group(
                    db, brand_id, group.id, payload, actor, import_id=import_id
                )
                updated += 1
            else:
                if not name:
                    raise ValueError("New reporting group rows require a name")

                payload = ReportingGroupCreate(name=name)
                await reporting_group_service.create_reporting_group(
                    db, brand_id, payload, actor, import_id=import_id
                )
                created += 1
        except (ValueError, HTTPException) as exc:
            errors.append(ImportRowError(row_number=row_number, message=_error_message(exc)))

    return ImportSummary(import_id=import_id, created=created, updated=updated, errors=errors)


# ── Variants (Stage 22) ───────────────────────────────────────────────────────


async def import_variants(
    db: AsyncSession,
    brand_id: uuid.UUID,
    file_bytes: bytes,
    actor: User | SuperAdmin,
) -> ImportSummary:
    """
    Bulk update Variants from an uploaded XLSX sheet (see VARIANT_COLUMNS in export_service.py).

    Update-only: every row must carry a `ref` matching an existing variant.
    Creating a new variant via import is not supported — attribute assignment
    varies per brand and doesn't fit a fixed spreadsheet header, so new
    variants must still be created from the Product page.

    Args:
        db: Active database session.
        brand_id: Brand to import into.
        file_bytes: Raw bytes of the uploaded .xlsx file.
        actor: The authenticated user performing the import (for audit logging).

    Returns:
        ImportSummary: Updated count and any skipped-row errors (created is always 0).
    """
    headers, rows = parse_xlsx(file_bytes)
    import_id = uuid.uuid4()
    updated = 0
    errors: list[ImportRowError] = []

    for row_number, row in enumerate(rows, start=2):
        try:
            ref = _clean_str(row.get("ref"))
            if not ref:
                raise ValueError("Variant rows require a ref — creating a new variant via import is not supported")

            variant = await _find_variant_by_ref(db, brand_id, ref)
            if variant is None:
                raise ValueError(f"No variant found with ref '{ref}'")

            sku = _clean_str(row.get("sku")) if "sku" in headers else None
            display_name = _clean_str(row.get("display_name")) if "display_name" in headers else None
            price_cents = _parse_price_cents(row.get("price")) if "price" in headers else None
            is_active = _parse_bool(row.get("is_active")) if "is_active" in headers else None

            payload = VariantUpdate(sku=sku, price_cents=price_cents, display_name=display_name)
            await variant_service.update_variant(
                db, brand_id, variant.id, payload, actor, import_id=import_id
            )
            if is_active is not None:
                await variant_service.set_variant_active_state(
                    db, brand_id, variant.id, is_active, actor, import_id=import_id
                )
            updated += 1
        except (ValueError, HTTPException) as exc:
            errors.append(ImportRowError(row_number=row_number, message=_error_message(exc)))

    return ImportSummary(import_id=import_id, created=0, updated=updated, errors=errors)


# ── Combos (Stage 22) ─────────────────────────────────────────────────────────


async def import_combos(
    db: AsyncSession,
    brand_id: uuid.UUID,
    file_bytes: bytes,
    actor: User | SuperAdmin,
) -> ImportSummary:
    """
    Bulk import Combos from an uploaded XLSX sheet (see COMBO_COLUMNS in export_service.py).

    Args:
        db: Active database session.
        brand_id: Brand to import into.
        file_bytes: Raw bytes of the uploaded .xlsx file.
        actor: The authenticated user performing the import (for audit logging).

    Returns:
        ImportSummary: Created/updated counts and any skipped-row errors.
    """
    headers, rows = parse_xlsx(file_bytes)
    import_id = uuid.uuid4()
    created = 0
    updated = 0
    errors: list[ImportRowError] = []

    product_map = await _product_id_by_ref(db, brand_id) if "product_ref" in headers else {}

    for row_number, row in enumerate(rows, start=2):
        try:
            ref = _clean_str(row.get("ref"))

            name = _clean_str(row.get("name")) if "name" in headers else None
            display_name = _clean_str(row.get("display_name")) if "display_name" in headers else None
            min_selections = _parse_int(row.get("min_selections")) if "min_selections" in headers else None
            max_selections = _parse_int(row.get("max_selections")) if "max_selections" in headers else None
            is_required = _parse_bool(row.get("is_required")) if "is_required" in headers else None
            display_order = _parse_int(row.get("display_order")) if "display_order" in headers else None
            is_active = _parse_bool(row.get("is_active")) if "is_active" in headers else None

            if ref:
                group = await _find_combo_group_by_ref(db, brand_id, ref)
                if group is None:
                    raise ValueError(f"No combo found with ref '{ref}'")

                payload = ComboGroupUpdate(
                    name=name,
                    display_name=display_name,
                    min_selections=min_selections,
                    max_selections=max_selections,
                    is_required=is_required,
                    display_order=display_order,
                )
                await combo_service.update_combo_group(
                    db, brand_id, group.id, payload, actor, import_id=import_id
                )
                if is_active is not None:
                    await combo_service.set_combo_group_active_state(
                        db, brand_id, group.id, is_active, actor, import_id=import_id
                    )
                updated += 1
            else:
                product_ref = _clean_str(row.get("product_ref")) if "product_ref" in headers else None
                product_id = product_map.get(product_ref) if product_ref else None
                if product_id is None:
                    raise ValueError(f"Unknown product_ref '{product_ref}'")
                if not name:
                    raise ValueError("New combo rows require a name")

                payload = ComboGroupCreate(
                    name=name,
                    display_name=display_name,
                    min_selections=min_selections if min_selections is not None else 1,
                    max_selections=max_selections if max_selections is not None else 1,
                    is_required=is_required if is_required is not None else True,
                    display_order=display_order if display_order is not None else 0,
                )
                await combo_service.create_combo_group(
                    db, brand_id, product_id, payload, actor, import_id=import_id
                )
                created += 1
        except (ValueError, HTTPException) as exc:
            errors.append(ImportRowError(row_number=row_number, message=_error_message(exc)))

    return ImportSummary(import_id=import_id, created=created, updated=updated, errors=errors)

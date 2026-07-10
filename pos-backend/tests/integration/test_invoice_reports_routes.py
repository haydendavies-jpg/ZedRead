"""Integration tests for invoice reporting routes (Stage 21).

Covers: filtered list (status/date/amount/site), XLSX export, detail view,
PDF export, and the change-log panel — including the refund fix that logs
the event against the *original* invoice's entity_id, not just the new
refund invoice's.
"""

import uuid
from io import BytesIO

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.audit_log import AuditLog


# ── Helpers ───────────────────────────────────────────────────────────────────


async def _create_paid_invoice(client: AsyncClient, headers: dict, product_id: str, amount_cents: int) -> str:
    """Create an invoice, add a line item, and pay it. Returns the invoice id."""
    inv_resp = await client.post("/invoices", json={}, headers=headers)
    assert inv_resp.status_code == 201, inv_resp.text
    invoice_id = inv_resp.json()["id"]

    li_resp = await client.post(
        f"/invoices/{invoice_id}/line-items",
        json={"product_id": product_id, "quantity": 1},
        headers=headers,
    )
    assert li_resp.status_code == 201, li_resp.text

    pay_resp = await client.post(
        f"/invoices/{invoice_id}/pay",
        json={"method": "cash", "amount_cents": amount_cents},
        headers=headers,
    )
    assert pay_resp.status_code == 200, pay_resp.text
    return invoice_id


# ── List (filters) ────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_list_invoice_reports_happy_path(
    client: AsyncClient, pos_auth_headers: dict, test_product
) -> None:
    """GET /invoice-reports returns the site's invoices, most recent first."""
    invoice_id = await _create_paid_invoice(client, pos_auth_headers, str(test_product.id), 1500)

    resp = await client.get("/invoice-reports", headers=pos_auth_headers)
    assert resp.status_code == 200
    ids = [row["id"] for row in resp.json()]
    assert invoice_id in ids


@pytest.mark.asyncio
async def test_list_invoice_reports_status_filter(
    client: AsyncClient, pos_auth_headers: dict, test_product
) -> None:
    """?status=paid excludes draft invoices from the filtered list."""
    await client.post("/invoices", json={}, headers=pos_auth_headers)  # draft, no line items
    paid_id = await _create_paid_invoice(client, pos_auth_headers, str(test_product.id), 1500)

    resp = await client.get("/invoice-reports", params={"status": "paid"}, headers=pos_auth_headers)
    assert resp.status_code == 200
    rows = resp.json()
    assert all(r["status"] == "paid" for r in rows)
    assert paid_id in [r["id"] for r in rows]


@pytest.mark.asyncio
async def test_list_invoice_reports_amount_range_filter(
    client: AsyncClient, pos_auth_headers: dict, test_product
) -> None:
    """min_amount_cents/max_amount_cents narrow the list to invoices in that total range."""
    low_id = await _create_paid_invoice(client, pos_auth_headers, str(test_product.id), 1500)

    resp = await client.get(
        "/invoice-reports",
        params={"min_amount_cents": 100000, "max_amount_cents": 200000},
        headers=pos_auth_headers,
    )
    assert resp.status_code == 200
    ids = [row["id"] for row in resp.json()]
    assert low_id not in ids


@pytest.mark.asyncio
async def test_list_invoice_reports_wrong_site_returns_403(
    client: AsyncClient, pos_auth_headers: dict
) -> None:
    """A POS-scoped caller requesting a different site_id is rejected."""
    wrong_site_id = str(uuid.uuid4())
    resp = await client.get(
        "/invoice-reports", params={"site_id": wrong_site_id}, headers=pos_auth_headers
    )
    assert resp.status_code == 403


@pytest.mark.asyncio
async def test_list_invoice_reports_no_auth_returns_403(client: AsyncClient) -> None:
    """No Authorization header returns 403 (FastAPI's HTTPBearer default for a missing header)."""
    resp = await client.get("/invoice-reports")
    assert resp.status_code == 403


# ── Detail view ───────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_get_invoice_detail_returns_line_items_and_payments(
    client: AsyncClient, pos_auth_headers: dict, test_product
) -> None:
    """GET /invoice-reports/{id} returns line items, tax breakdown, and payments."""
    invoice_id = await _create_paid_invoice(client, pos_auth_headers, str(test_product.id), 1500)

    resp = await client.get(f"/invoice-reports/{invoice_id}", headers=pos_auth_headers)
    assert resp.status_code == 200
    data = resp.json()
    assert data["id"] == invoice_id
    assert len(data["line_items"]) == 1
    assert data["line_items"][0]["product_name"] == test_product.name
    assert len(data["payments"]) == 1
    assert data["payments"][0]["amount_cents"] == 1500


@pytest.mark.asyncio
async def test_get_invoice_detail_not_found_returns_404(
    client: AsyncClient, pos_auth_headers: dict
) -> None:
    """A random invoice id within the caller's brand returns 404."""
    resp = await client.get(f"/invoice-reports/{uuid.uuid4()}", headers=pos_auth_headers)
    assert resp.status_code == 404


# ── Change log ────────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_change_log_includes_created_and_paid_actions(
    client: AsyncClient, pos_auth_headers: dict, test_product
) -> None:
    """The change-log panel surfaces invoice.created and invoice.paid entries."""
    invoice_id = await _create_paid_invoice(client, pos_auth_headers, str(test_product.id), 1500)

    resp = await client.get(f"/invoice-reports/{invoice_id}/change-log", headers=pos_auth_headers)
    assert resp.status_code == 200
    actions = [entry["action"] for entry in resp.json()]
    assert "invoice.created" in actions
    assert "invoice.paid" in actions


@pytest.mark.asyncio
async def test_change_log_includes_refund_event_on_original_invoice(
    client: AsyncClient, pos_auth_headers: dict, test_product, db: AsyncSession
) -> None:
    """
    Refunding an invoice must show up in the ORIGINAL invoice's change log,
    not only the new refund invoice's — this is the Stage 21 fix to
    create_refund(), which previously only logged against the refund's own id.
    """
    invoice_id = await _create_paid_invoice(client, pos_auth_headers, str(test_product.id), 1500)

    refund_resp = await client.post(
        f"/invoices/{invoice_id}/refund", json={"reason": "test"}, headers=pos_auth_headers
    )
    assert refund_resp.status_code == 201

    resp = await client.get(f"/invoice-reports/{invoice_id}/change-log", headers=pos_auth_headers)
    assert resp.status_code == 200
    actions = [entry["action"] for entry in resp.json()]
    assert "invoice.refunded" in actions

    # Confirm the underlying audit row is scoped to the original invoice's own id
    result = await db.execute(
        select(AuditLog).where(
            AuditLog.entity_type == "invoice",
            AuditLog.entity_id == invoice_id,
            AuditLog.action == "invoice.refunded",
        )
    )
    assert result.scalar_one_or_none() is not None


# ── XLSX export ───────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_export_invoices_report_returns_xlsx(
    client: AsyncClient, pos_auth_headers: dict, test_product
) -> None:
    """GET /invoice-reports/export streams a re-readable .xlsx workbook."""
    await _create_paid_invoice(client, pos_auth_headers, str(test_product.id), 1500)

    resp = await client.get("/invoice-reports/export", headers=pos_auth_headers)
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    header_row = [cell.value for cell in ws[1]]
    assert header_row == [
        "id", "site", "invoice_type", "status", "created_at",
        "subtotal", "tax", "discount", "total", "is_refunded", "voided_at", "paid_at",
    ]
    assert ws.max_row >= 2  # header + at least one invoice row


# ── PDF export ────────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_export_invoice_pdf_returns_pdf_bytes(
    client: AsyncClient, pos_auth_headers: dict, test_product
) -> None:
    """GET /invoice-reports/{id}/pdf streams a valid PDF document."""
    invoice_id = await _create_paid_invoice(client, pos_auth_headers, str(test_product.id), 1500)

    resp = await client.get(f"/invoice-reports/{invoice_id}/pdf", headers=pos_auth_headers)
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"
    assert resp.content.startswith(b"%PDF")

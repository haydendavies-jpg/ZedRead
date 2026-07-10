"""Integration tests for the Stage 19 Products bulk import/export routes.

Covers:
1. Happy path — template download, full export, create-by-import, update-by-import
2. Auth failure — no token returns 403
3. Invalid input — malformed .xlsx returns 422
4. Business rules — unknown ref/category are skipped as row errors, not a hard failure
5. Audit log — PRODUCT_CREATED/PRODUCT_UPDATED rows carry the shared import_id
"""

import io

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from app.constants.audit_actions import PRODUCT_CREATED, PRODUCT_UPDATED
from app.models.audit_log import AuditLog
from app.models.product import Product

pytestmark = pytest.mark.asyncio


def _xlsx_bytes(headers: list[str], rows: list[list]) -> bytes:
    """Build minimal XLSX bytes with a header row and the given data rows."""
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _upload(client, headers, file_bytes):
    return client.post(
        "/products/import",
        files={"file": ("import.xlsx", io.BytesIO(file_bytes), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )


# ── Template / export ─────────────────────────────────────────────────────────


async def test_export_template_returns_xlsx(client, mgmt_auth_headers, test_product):
    """GET /products/export/template returns a downloadable workbook with the expected headers."""
    response = await client.get("/products/export/template", headers=mgmt_auth_headers)

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.openxmlformats")
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    header_row = [c.value for c in ws[1]]
    assert header_row == ["ref", "name", "category", "description", "print_name", "price", "is_taxable", "is_open_item", "display_order", "is_active"]


async def test_export_returns_current_catalog(client, mgmt_auth_headers, test_product):
    """GET /products/export includes the brand's existing product data."""
    response = await client.get("/products/export", headers=mgmt_auth_headers)

    assert response.status_code == 200
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    refs = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert test_product.ref in refs


async def test_export_no_token_returns_403(client):
    """GET /products/export with no Authorization header returns 403."""
    response = await client.get("/products/export")
    assert response.status_code == 403


# ── Import: create ────────────────────────────────────────────────────────────


async def test_import_creates_new_product(client, db, mgmt_auth_headers, test_brand, test_product):
    """A row with a blank ref creates a new product."""
    from app.models.category import Category

    cat_result = await db.execute(select(Category).where(Category.brand_id == test_brand.id))
    category = cat_result.scalars().first()

    file_bytes = _xlsx_bytes(
        ["ref", "name", "category", "price", "is_taxable"],
        [["", "Imported Burger", category.name, "12.50", "TRUE"]],
    )

    response = await _upload(client, mgmt_auth_headers, file_bytes)

    assert response.status_code == 200
    body = response.json()
    assert body["created"] == 1
    assert body["updated"] == 0
    assert body["errors"] == []

    result = await db.execute(select(Product).where(Product.name == "Imported Burger"))
    product = result.scalar_one()
    assert product.base_price_cents == 1250
    assert product.category_id == category.id


async def test_import_create_writes_audit_log_with_import_id(client, db, mgmt_auth_headers, test_brand, test_product):
    """A created-by-import row writes a PRODUCT_CREATED audit row carrying the batch import_id."""
    from app.models.category import Category

    cat_result = await db.execute(select(Category).where(Category.brand_id == test_brand.id))
    category = cat_result.scalars().first()

    file_bytes = _xlsx_bytes(["ref", "name", "category", "price"], [["", "Audit Burger", category.name, "5.00"]])
    response = await _upload(client, mgmt_auth_headers, file_bytes)
    import_id = response.json()["import_id"]

    result = await db.execute(select(Product).where(Product.name == "Audit Burger"))
    product = result.scalar_one()

    audit_result = await db.execute(
        select(AuditLog).where(AuditLog.entity_id == str(product.id), AuditLog.action == PRODUCT_CREATED)
    )
    row = audit_result.scalar_one()
    assert row.after_state["import_id"] == import_id


async def test_import_missing_required_field_is_row_error(client, mgmt_auth_headers):
    """A new-row without a resolvable category is skipped and reported, not a hard failure."""
    file_bytes = _xlsx_bytes(["ref", "name", "category", "price"], [["", "No Category Product", "Nonexistent", "5.00"]])

    response = await _upload(client, mgmt_auth_headers, file_bytes)

    assert response.status_code == 200
    body = response.json()
    assert body["created"] == 0
    assert len(body["errors"]) == 1
    assert body["errors"][0]["row_number"] == 2
    assert "Nonexistent" in body["errors"][0]["message"]


# ── Import: update (partial-update semantics) ────────────────────────────────


async def test_import_updates_existing_product_by_ref(client, db, mgmt_auth_headers, test_product):
    """A row matched by ref updates only the columns present in the header."""
    file_bytes = _xlsx_bytes(["ref", "name"], [[test_product.ref, "Renamed Via Import"]])

    response = await _upload(client, mgmt_auth_headers, file_bytes)

    assert response.status_code == 200
    body = response.json()
    assert body["updated"] == 1
    assert body["created"] == 0

    await db.refresh(test_product)
    assert test_product.name == "Renamed Via Import"
    # Price wasn't in the header — must be untouched
    assert test_product.base_price_cents == 1500


async def test_import_update_writes_audit_log(client, db, mgmt_auth_headers, test_product):
    """An updated-by-import row writes a PRODUCT_UPDATED audit row."""
    file_bytes = _xlsx_bytes(["ref", "name"], [[test_product.ref, "Audited Rename"]])
    await _upload(client, mgmt_auth_headers, file_bytes)

    result = await db.execute(
        select(AuditLog).where(AuditLog.entity_id == str(test_product.id), AuditLog.action == PRODUCT_UPDATED)
    )
    row = result.scalar_one()
    assert row.after_state["import_id"] is not None


async def test_import_unknown_ref_is_row_error(client, mgmt_auth_headers):
    """A ref that doesn't match any existing product is a row error, not a new record."""
    file_bytes = _xlsx_bytes(["ref", "name"], [["PRD-999999", "Ghost Product"]])

    response = await _upload(client, mgmt_auth_headers, file_bytes)

    assert response.status_code == 200
    body = response.json()
    assert body["updated"] == 0
    assert body["created"] == 0
    assert "PRD-999999" in body["errors"][0]["message"]


async def test_import_sets_is_active_false(client, db, mgmt_auth_headers, test_product):
    """An is_active column set to FALSE soft-deletes the matched product."""
    file_bytes = _xlsx_bytes(["ref", "is_active"], [[test_product.ref, "FALSE"]])

    response = await _upload(client, mgmt_auth_headers, file_bytes)

    assert response.status_code == 200
    await db.refresh(test_product)
    assert test_product.is_active is False


async def test_import_no_token_returns_403(client):
    """POST /products/import with no Authorization header returns 403."""
    file_bytes = _xlsx_bytes(["ref", "name"], [])
    response = await _upload(client, {}, file_bytes)
    assert response.status_code == 403


async def test_import_invalid_file_returns_422(client, mgmt_auth_headers):
    """Uploading a non-.xlsx file returns 422."""
    response = await client.post(
        "/products/import",
        files={"file": ("import.xlsx", io.BytesIO(b"not a real workbook"), "application/octet-stream")},
        headers=mgmt_auth_headers,
    )
    assert response.status_code == 422

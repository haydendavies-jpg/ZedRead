"""Integration tests for the Stage 19 Reporting Groups bulk import/export routes.

Covers:
1. Happy path — template download, full export, create-by-import, update-by-import
2. Auth failure — no token returns 403; POS token returns 403 on import
3. Invalid input — malformed .xlsx returns 422
4. Business rules — unknown ref is a row error; renaming the system default group
   via import is rejected same as the direct PATCH route
5. Audit log — REPORTING_GROUP_CREATED/REPORTING_GROUP_UPDATED rows carry the shared import_id
"""

import io

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from app.constants.audit_actions import REPORTING_GROUP_CREATED, REPORTING_GROUP_UPDATED
from app.models.audit_log import AuditLog
from app.models.reporting_group import ReportingGroup

pytestmark = pytest.mark.asyncio


def _xlsx_bytes(headers: list[str], rows: list[list]) -> bytes:
    """Build minimal XLSX bytes with a header row and the given data rows."""
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _upload(client, headers, file_bytes):
    return client.post(
        "/reporting-groups/import",
        files={"file": ("import.xlsx", io.BytesIO(file_bytes), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )


# ── Template / export ─────────────────────────────────────────────────────────


async def test_export_template_returns_xlsx(client, mgmt_auth_headers):
    """GET /reporting-groups/export/template returns a downloadable workbook with the expected headers."""
    response = await client.get("/reporting-groups/export/template", headers=mgmt_auth_headers)

    assert response.status_code == 200
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["ref", "name"]


async def test_export_returns_current_groups(client, mgmt_auth_headers, test_reporting_group):
    """GET /reporting-groups/export includes the brand's existing reporting groups."""
    response = await client.get("/reporting-groups/export", headers=mgmt_auth_headers)

    assert response.status_code == 200
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    refs = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert test_reporting_group.ref in refs


async def test_export_no_token_returns_403(client):
    """GET /reporting-groups/export with no Authorization header returns 403."""
    response = await client.get("/reporting-groups/export")
    assert response.status_code == 403


# ── Import: create ────────────────────────────────────────────────────────────


async def test_import_creates_new_reporting_group(client, db, mgmt_auth_headers, test_brand):
    """A row with a blank ref creates a new reporting group."""
    file_bytes = _xlsx_bytes(["ref", "name"], [["", "Imported Group"]])

    response = await _upload(client, mgmt_auth_headers, file_bytes)

    assert response.status_code == 200
    body = response.json()
    assert body["created"] == 1
    assert body["errors"] == []

    result = await db.execute(select(ReportingGroup).where(ReportingGroup.name == "Imported Group"))
    group = result.scalar_one()
    assert group.brand_id == test_brand.id
    assert group.is_default is False


async def test_import_create_writes_audit_log_with_import_id(client, db, mgmt_auth_headers):
    """A created-by-import row writes a REPORTING_GROUP_CREATED audit row carrying the batch import_id."""
    file_bytes = _xlsx_bytes(["ref", "name"], [["", "Audited Group"]])
    response = await _upload(client, mgmt_auth_headers, file_bytes)
    import_id = response.json()["import_id"]

    result = await db.execute(select(ReportingGroup).where(ReportingGroup.name == "Audited Group"))
    group = result.scalar_one()

    audit_result = await db.execute(
        select(AuditLog).where(AuditLog.entity_id == str(group.id), AuditLog.action == REPORTING_GROUP_CREATED)
    )
    row = audit_result.scalar_one()
    assert row.after_state["import_id"] == import_id


async def test_import_new_row_missing_name_is_row_error(client, mgmt_auth_headers):
    """A new row (no ref) with a blank name is reported as a row error, not silently skipped.

    A third, unrecognised column carries a value so the row isn't dropped by
    parse_xlsx's fully-blank-row filter before validation ever runs.
    """
    file_bytes = _xlsx_bytes(["ref", "name", "note"], [["", "", "placeholder"]])

    response = await _upload(client, mgmt_auth_headers, file_bytes)

    assert response.status_code == 200
    body = response.json()
    assert body["created"] == 0
    assert len(body["errors"]) == 1


# ── Import: update ────────────────────────────────────────────────────────────


async def test_import_updates_existing_group_by_ref(client, db, mgmt_auth_headers):
    """A row matched by ref renames a non-system reporting group."""
    create_resp = await client.post("/reporting-groups", json={"name": "Original"}, headers=mgmt_auth_headers)
    group_id = create_resp.json()["id"]
    group_ref = create_resp.json()["ref"]

    file_bytes = _xlsx_bytes(["ref", "name"], [[group_ref, "Renamed Via Import"]])
    response = await _upload(client, mgmt_auth_headers, file_bytes)

    assert response.status_code == 200
    assert response.json()["updated"] == 1

    result = await db.execute(select(ReportingGroup).where(ReportingGroup.id == group_id))
    group = result.scalar_one()
    assert group.name == "Renamed Via Import"


async def test_import_update_writes_audit_log(client, db, mgmt_auth_headers):
    """An updated-by-import row writes a REPORTING_GROUP_UPDATED audit row."""
    create_resp = await client.post("/reporting-groups", json={"name": "Original2"}, headers=mgmt_auth_headers)
    group_id = create_resp.json()["id"]
    group_ref = create_resp.json()["ref"]

    file_bytes = _xlsx_bytes(["ref", "name"], [[group_ref, "Renamed2"]])
    await _upload(client, mgmt_auth_headers, file_bytes)

    result = await db.execute(
        select(AuditLog).where(AuditLog.entity_id == group_id, AuditLog.action == REPORTING_GROUP_UPDATED)
    )
    row = result.scalar_one()
    assert row.after_state["import_id"] is not None


async def test_import_unknown_ref_is_row_error(client, mgmt_auth_headers):
    """A ref that doesn't match any existing reporting group is a row error."""
    file_bytes = _xlsx_bytes(["ref", "name"], [["RPG-999999", "Ghost Group"]])

    response = await _upload(client, mgmt_auth_headers, file_bytes)

    assert response.status_code == 200
    body = response.json()
    assert body["updated"] == 0
    assert "RPG-999999" in body["errors"][0]["message"]


async def test_import_rename_default_group_is_row_error(client, db, mgmt_auth_headers, test_reporting_group):
    """Renaming the brand's system default reporting group via import is rejected, not applied."""
    file_bytes = _xlsx_bytes(["ref", "name"], [[test_reporting_group.ref, "Hacked Name"]])

    response = await _upload(client, mgmt_auth_headers, file_bytes)

    assert response.status_code == 200
    body = response.json()
    assert body["updated"] == 0
    assert len(body["errors"]) == 1

    await db.refresh(test_reporting_group)
    assert test_reporting_group.name != "Hacked Name"


# ── Auth ──────────────────────────────────────────────────────────────────────


async def test_import_pos_token_returns_403(client, pos_auth_headers):
    """POST /reporting-groups/import with a POS terminal token returns 403."""
    file_bytes = _xlsx_bytes(["ref", "name"], [])
    response = await _upload(client, pos_auth_headers, file_bytes)
    assert response.status_code == 403


async def test_import_invalid_file_returns_422(client, mgmt_auth_headers):
    """Uploading a non-.xlsx file returns 422."""
    response = await client.post(
        "/reporting-groups/import",
        files={"file": ("import.xlsx", io.BytesIO(b"not a real workbook"), "application/octet-stream")},
        headers=mgmt_auth_headers,
    )
    assert response.status_code == 422

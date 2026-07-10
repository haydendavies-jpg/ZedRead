"""Integration tests for the Stage 19 Categories bulk import/export routes.

Covers:
1. Happy path — template download, full export, create-by-import, update-by-import
2. Auth failure — no token returns 403; POS token returns 403 on import
3. Invalid input — malformed .xlsx returns 422
4. Business rules — unknown ref/reporting group are row errors; system category
   rename via import is rejected same as the direct PATCH route
5. Audit log — CATEGORY_CREATED/CATEGORY_UPDATED rows carry the shared import_id
"""

import io

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from app.constants.audit_actions import CATEGORY_CREATED, CATEGORY_UPDATED
from app.models.audit_log import AuditLog
from app.models.category import Category

pytestmark = pytest.mark.asyncio


def _xlsx_bytes(headers: list[str], rows: list[list]) -> bytes:
    """Build minimal XLSX bytes with a header row and the given data rows."""
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _upload(client, headers, file_bytes):
    return client.post(
        "/categories/import",
        files={"file": ("import.xlsx", io.BytesIO(file_bytes), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )


# ── Template / export ─────────────────────────────────────────────────────────


async def test_export_template_returns_xlsx(client, mgmt_auth_headers, test_reporting_group):
    """GET /categories/export/template returns a downloadable workbook with the expected headers."""
    response = await client.get("/categories/export/template", headers=mgmt_auth_headers)

    assert response.status_code == 200
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["ref", "name", "reporting_group", "display_order", "is_active"]


async def test_export_returns_current_categories(client, mgmt_auth_headers, test_product):
    """GET /categories/export includes the brand's existing category data."""
    response = await client.get("/categories/export", headers=mgmt_auth_headers)

    assert response.status_code == 200
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    names = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
    assert "Uncategorised" in names


async def test_export_no_token_returns_403(client):
    """GET /categories/export with no Authorization header returns 403."""
    response = await client.get("/categories/export")
    assert response.status_code == 403


# ── Import: create ────────────────────────────────────────────────────────────


async def test_import_creates_new_category(client, db, mgmt_auth_headers, test_brand, test_reporting_group):
    """A row with a blank ref creates a new category, resolving reporting_group by name."""
    file_bytes = _xlsx_bytes(
        ["ref", "name", "reporting_group", "display_order"],
        [["", "Imported Mains", test_reporting_group.name, "1"]],
    )

    response = await _upload(client, mgmt_auth_headers, file_bytes)

    assert response.status_code == 200
    body = response.json()
    assert body["created"] == 1
    assert body["errors"] == []

    result = await db.execute(select(Category).where(Category.name == "Imported Mains"))
    category = result.scalar_one()
    assert category.reporting_group_id == test_reporting_group.id
    assert category.display_order == 1


async def test_import_create_writes_audit_log_with_import_id(client, db, mgmt_auth_headers, test_reporting_group):
    """A created-by-import row writes a CATEGORY_CREATED audit row carrying the batch import_id."""
    file_bytes = _xlsx_bytes(["ref", "name", "reporting_group"], [["", "Audit Category", test_reporting_group.name]])
    response = await _upload(client, mgmt_auth_headers, file_bytes)
    import_id = response.json()["import_id"]

    result = await db.execute(select(Category).where(Category.name == "Audit Category"))
    category = result.scalar_one()

    audit_result = await db.execute(
        select(AuditLog).where(AuditLog.entity_id == str(category.id), AuditLog.action == CATEGORY_CREATED)
    )
    row = audit_result.scalar_one()
    assert row.after_state["import_id"] == import_id


async def test_import_unknown_reporting_group_is_row_error(client, mgmt_auth_headers):
    """A reporting_group name that doesn't resolve is a row error, not silently defaulted."""
    file_bytes = _xlsx_bytes(["ref", "name", "reporting_group"], [["", "Orphan Category", "Nonexistent Group"]])

    response = await _upload(client, mgmt_auth_headers, file_bytes)

    assert response.status_code == 200
    body = response.json()
    assert body["created"] == 0
    assert "Nonexistent Group" in body["errors"][0]["message"]


async def test_import_new_category_without_reporting_group_column_uses_default(
    client, db, mgmt_auth_headers, test_reporting_group
):
    """Omitting the reporting_group column entirely still auto-assigns the brand default (create_category's existing behaviour)."""
    file_bytes = _xlsx_bytes(["ref", "name"], [["", "Defaulted Category"]])

    response = await _upload(client, mgmt_auth_headers, file_bytes)

    assert response.status_code == 200
    assert response.json()["created"] == 1

    result = await db.execute(select(Category).where(Category.name == "Defaulted Category"))
    category = result.scalar_one()
    assert category.reporting_group_id == test_reporting_group.id


# ── Import: update (partial-update semantics) ────────────────────────────────


async def test_import_updates_existing_category_by_ref(client, db, mgmt_auth_headers, test_reporting_group):
    """A row matched by ref renames only the columns present in the header."""
    create_resp = await client.post(
        "/categories", json={"name": "Original", "brand_id": str(test_reporting_group.brand_id)}, headers=mgmt_auth_headers
    )
    category_id = create_resp.json()["id"]
    category_ref = create_resp.json()["ref"]

    file_bytes = _xlsx_bytes(["ref", "name"], [[category_ref, "Renamed Via Import"]])
    response = await _upload(client, mgmt_auth_headers, file_bytes)

    assert response.status_code == 200
    assert response.json()["updated"] == 1

    result = await db.execute(select(Category).where(Category.id == category_id))
    category = result.scalar_one()
    assert category.name == "Renamed Via Import"


async def test_import_update_writes_audit_log(client, db, mgmt_auth_headers, test_reporting_group):
    """An updated-by-import row writes a CATEGORY_UPDATED audit row."""
    create_resp = await client.post(
        "/categories", json={"name": "Original2", "brand_id": str(test_reporting_group.brand_id)}, headers=mgmt_auth_headers
    )
    category_id = create_resp.json()["id"]
    category_ref = create_resp.json()["ref"]

    file_bytes = _xlsx_bytes(["ref", "name"], [[category_ref, "Renamed2"]])
    await _upload(client, mgmt_auth_headers, file_bytes)

    result = await db.execute(
        select(AuditLog).where(AuditLog.entity_id == category_id, AuditLog.action == CATEGORY_UPDATED)
    )
    row = result.scalar_one()
    assert row.after_state["import_id"] is not None


async def test_import_unknown_ref_is_row_error(client, mgmt_auth_headers):
    """A ref that doesn't match any existing category is a row error."""
    file_bytes = _xlsx_bytes(["ref", "name"], [["CAT-999999", "Ghost Category"]])

    response = await _upload(client, mgmt_auth_headers, file_bytes)

    assert response.status_code == 200
    body = response.json()
    assert body["updated"] == 0
    assert "CAT-999999" in body["errors"][0]["message"]


async def test_import_rename_system_category_is_row_error(client, db, mgmt_auth_headers, test_product):
    """Renaming the system 'Uncategorised' category via import is rejected, not applied."""
    result = await db.execute(select(Category).where(Category.is_system == True))  # noqa: E712
    system_category = result.scalars().first()

    file_bytes = _xlsx_bytes(["ref", "name"], [[system_category.ref, "Hacked Name"]])
    response = await _upload(client, mgmt_auth_headers, file_bytes)

    assert response.status_code == 200
    body = response.json()
    assert body["updated"] == 0
    assert len(body["errors"]) == 1

    await db.refresh(system_category)
    assert system_category.name != "Hacked Name"


# ── Auth ──────────────────────────────────────────────────────────────────────


async def test_import_pos_token_returns_403(client, pos_auth_headers):
    """POST /categories/import with a POS terminal token returns 403."""
    file_bytes = _xlsx_bytes(["ref", "name"], [])
    response = await _upload(client, pos_auth_headers, file_bytes)
    assert response.status_code == 403


async def test_import_invalid_file_returns_422(client, mgmt_auth_headers):
    """Uploading a non-.xlsx file returns 422."""
    response = await client.post(
        "/categories/import",
        files={"file": ("import.xlsx", io.BytesIO(b"not a real workbook"), "application/octet-stream")},
        headers=mgmt_auth_headers,
    )
    assert response.status_code == 422

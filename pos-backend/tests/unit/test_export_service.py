"""Unit tests for the Stage 19 XLSX-building helpers in export_service.py.

No database required — these exercise the pure workbook-assembly logic.
Query-backed functions (export_products, build_products_export, etc.) are
covered by the integration route tests instead, since they need real rows.
"""

from openpyxl import load_workbook

from app.services.export_service import (
    CATEGORY_COLUMNS,
    PRODUCT_COLUMNS,
    REPORTING_GROUP_COLUMNS,
    _add_name_dropdown,
    _cents_to_dollars_str,
    _rows_to_workbook,
    workbook_to_bytes,
)


def test_cents_to_dollars_str_formats_two_decimals():
    assert _cents_to_dollars_str(999) == "9.99"
    assert _cents_to_dollars_str(1000) == "10.00"
    assert _cents_to_dollars_str(5) == "0.05"


def test_rows_to_workbook_writes_header_and_rows():
    """The header row matches the given order and data rows follow it."""
    wb = _rows_to_workbook(["ref", "name"], [{"ref": "CAT-000001", "name": "Mains"}], "Categories")
    ws = wb.active
    assert ws.title == "Categories"
    assert [c.value for c in ws[1]] == ["ref", "name"]
    assert [c.value for c in ws[2]] == ["CAT-000001", "Mains"]
    assert ws.freeze_panes == "A2"


def test_rows_to_workbook_missing_key_in_row_becomes_blank():
    """A data row dict missing a header key writes an empty cell, not an error."""
    wb = _rows_to_workbook(["ref", "name"], [{"name": "Mains"}], "Categories")
    ws = wb.active
    assert [c.value for c in ws[2]] == ["", "Mains"]


def test_add_name_dropdown_creates_hidden_helper_sheet():
    """Dropdown options are written to a hidden sheet and referenced by a data validation."""
    wb = _rows_to_workbook(["ref", "category"], [], "Products")
    _add_name_dropdown(wb, "Products", 2, ["Mains", "Drinks"])

    helper_titles = [t for t in wb.sheetnames if t != "Products"]
    assert len(helper_titles) == 1
    helper = wb[helper_titles[0]]
    assert helper.sheet_state == "hidden"
    assert [helper.cell(row=i, column=2).value for i in (1, 2)] == ["Mains", "Drinks"]

    ws = wb["Products"]
    assert len(ws.data_validations.dataValidation) == 1


def test_add_name_dropdown_noop_when_no_options():
    """No options means no dropdown/helper sheet is created."""
    wb = _rows_to_workbook(["ref", "category"], [], "Products")
    _add_name_dropdown(wb, "Products", 2, [])
    assert wb.sheetnames == ["Products"]


def test_workbook_to_bytes_roundtrips_through_openpyxl():
    """Serialised bytes can be re-parsed and contain the same header/data."""
    import io

    wb = _rows_to_workbook(["ref", "name"], [{"ref": "RPG-000001", "name": "Default"}], "Reporting Groups")
    raw = workbook_to_bytes(wb)
    reloaded = load_workbook(io.BytesIO(raw))
    ws = reloaded.active
    assert [c.value for c in ws[1]] == ["ref", "name"]
    assert [c.value for c in ws[2]] == ["RPG-000001", "Default"]


def test_column_orders_match_import_contract():
    """Column lists are the shared header contract with import_service.py's `in headers` checks."""
    assert PRODUCT_COLUMNS[0] == "ref"
    assert "category" in PRODUCT_COLUMNS
    assert CATEGORY_COLUMNS[0] == "ref"
    assert "reporting_group" in CATEGORY_COLUMNS
    assert REPORTING_GROUP_COLUMNS == ["ref", "name"]
